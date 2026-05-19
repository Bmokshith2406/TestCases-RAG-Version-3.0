import pytest
import asyncio
from unittest.mock import AsyncMock, MagicMock, patch
from typing import Generator
from datetime import datetime, timezone
import uuid

from fastapi.testclient import TestClient
from motor.motor_asyncio import AsyncIOMotorClient, AsyncIOMotorDatabase

# Import app components
from app.main import app
from app.core.config import Settings, get_settings
from app.core.security import get_current_user, get_optional_current_user
from app.models.schemas import SearchRequest, SearchResultItem, SearchResponse
from app.db.mongo import get_db, get_client


# ========================================================================
# SETTINGS FIXTURE
# ========================================================================

@pytest.fixture(scope="session")
def mock_settings() -> Settings:
    """Mock settings for testing."""
    settings = Settings()
    settings.MONGO_CONNECTION_STRING = "mongodb://test:test@localhost:27017"
    settings.GOOGLE_API_KEY = "test-api-key"
    settings.OPENAI_API_KEY = ""
    settings.ANTHROPIC_API_KEY = ""
    settings.LOCAL_LLM_API_KEY = ""
    settings.LLM_USE_GEMINI = True
    settings.LLM_USE_OPENAI = False
    settings.LLM_USE_ANTHROPIC = False
    settings.LLM_USE_LOCAL = False
    settings.CACHE_ENABLED = True
    settings.CACHE_TTL_SECONDS = 300
    settings.CACHE_MAX_SIZE = 1000
    settings.GEMINI_RERANK_ENABLED = True
    settings.QUERY_EXPANSION_ENABLED = True
    settings.EMBEDDING_MODEL_NAME = "all-MiniLM-L6-v2"
    settings.EMBEDDING_USE_MINILM_384 = True
    settings.EMBEDDING_USE_MPNET_768 = False
    settings.EMBEDDING_USE_BGE_LARGE_1024 = False
    settings.VECTOR_INDEX_NAME = "vector_index"
    settings.EMBEDDING_DIMENSIONS = 384
    settings.CANDIDATES_TO_RETRIEVE = 15
    settings.FINAL_RESULTS = 5
    settings.TOP_K = 3
    settings.MAX_QUERY_LENGTH = 5000
    return settings


@pytest.fixture
def mock_get_settings(mock_settings):
    """Override get_settings in tests."""
    with patch("app.core.config.get_settings", return_value=mock_settings):
        yield mock_settings


# ========================================================================
# DATABASE FIXTURES
# ========================================================================

@pytest.fixture
def mock_mongo_client():
    """Mock MongoDB client."""
    client = AsyncMock(spec=AsyncIOMotorClient)
    return client


@pytest.fixture
def mock_mongo_db():
    """Mock MongoDB database instance."""
    db = AsyncMock(spec=AsyncIOMotorDatabase)
    return db


@pytest.fixture
def mock_testcase_collection():
    """Mock test case collection."""
    collection = AsyncMock()
    collection.find_one = AsyncMock()
    collection.find = AsyncMock()
    collection.insert_one = AsyncMock()
    collection.insert_many = AsyncMock()
    collection.update_one = AsyncMock()
    collection.delete_one = AsyncMock()
    collection.aggregate = AsyncMock()
    collection.create_index = AsyncMock()
    collection.count_documents = AsyncMock(return_value=0)
    return collection


@pytest.fixture
def mock_scripts_collection():
    """Mock Playwright scripts collection."""
    collection = AsyncMock()
    collection.find_one = AsyncMock()
    collection.insert_one = AsyncMock()
    collection.insert_many = AsyncMock()
    collection.update_one = AsyncMock()
    collection.delete_one = AsyncMock()
    return collection


# ========================================================================
# FASTAPI TEST CLIENT FIXTURE
# ========================================================================

@pytest.fixture
def test_client() -> TestClient:
    """FastAPI test client."""
    test_user = {
        "id": "test-user-id",
        "username": "test-admin",
        "role": "admin",
        "scopes": [
            "search:read",
            "cases:read",
            "scripts:read",
            "stats:read",
            "cases:write",
            "uploads:write",
            "admin:write",
            "users:write",
        ],
    }

    app.dependency_overrides[get_current_user] = lambda: test_user
    app.dependency_overrides[get_optional_current_user] = lambda: test_user

    client = TestClient(app)

    try:
        yield client
    finally:
        client.close()
        app.dependency_overrides.pop(get_current_user, None)
        app.dependency_overrides.pop(get_optional_current_user, None)


# ========================================================================
# EVENT LOOP FIXTURE (for async tests)
# ========================================================================

@pytest.fixture(scope="session")
def event_loop():
    """Create event loop for async tests."""
    policy = asyncio.get_event_loop_policy()
    loop = policy.new_event_loop()
    yield loop
    loop.close()


# ========================================================================
# SAMPLE DATA FIXTURES
# ========================================================================

@pytest.fixture
def sample_search_request() -> SearchRequest:
    """Sample search request."""
    return SearchRequest(
        query="login functionality test",
        feature="Authentication",
        ranking_variant="A",
    )


@pytest.fixture
def sample_search_result_items() -> list[SearchResultItem]:
    """Sample search result items."""
    return [
        SearchResultItem(
            id=str(uuid.uuid4()),
            probability=95.5,
            test_case_id="TC001",
            feature="Authentication",
            description="Test valid login with correct credentials",
            prerequisites="User account exists",
            steps="1. Navigate to login\n2. Enter credentials\n3. Click login",
            summary="Verify successful login with valid credentials",
            keywords=["login", "authentication", "valid", "credentials"],
            tags=["smoke", "critical"],
            priority="High",
            platform="Web",
            playwright_script_id=str(uuid.uuid4()),
        ),
        SearchResultItem(
            id=str(uuid.uuid4()),
            probability=87.2,
            test_case_id="TC002",
            feature="Authentication",
            description="Test invalid login attempt",
            prerequisites="User account exists",
            steps="1. Navigate to login\n2. Enter wrong password\n3. Click login",
            summary="Verify error message on invalid login",
            keywords=["login", "error", "invalid", "password"],
            tags=["smoke", "negative"],
            priority="High",
            platform="Web",
            playwright_script_id=str(uuid.uuid4()),
        ),
    ]


@pytest.fixture
def sample_testcase_doc() -> dict:
    """Sample test case document for MongoDB."""
    testcase_id = str(uuid.uuid4())
    script_id = str(uuid.uuid4())
    
    return {
        "_id": testcase_id,
        "Test Case ID": "TC-001",
        "Feature": "Login",
        "Test Case Description": "Test user login with valid credentials",
        "Pre-requisites": "User account must exist",
        "Steps": "1. Go to login page\n2. Enter credentials\n3. Click login",
        "TestCaseSummary": "Verify successful login with valid credentials",
        "TestCaseKeywords": ["login", "authentication", "valid"],
        "Tags": ["smoke", "critical"],
        "Priority": "High",
        "Platform": "Web",
        "desc_embedding": [0.1] * 384,
        "steps_embedding": [0.2] * 384,
        "summary_embedding": [0.3] * 384,
        "main_vector": [0.25] * 384,
        "playwright_script_id": script_id,
        "CreatedAt": datetime.now(timezone.utc),
        "Popularity": 0.85,
    }


@pytest.fixture
def sample_script_doc(sample_testcase_doc) -> dict:
    """Sample Playwright script document."""
    return {
        "_id": str(uuid.uuid4()),
        "testcase_id": sample_testcase_doc["Test Case ID"],
        "testcase_object_id": sample_testcase_doc["_id"],
        "script": """
        async function login() {
            await page.goto('http://localhost:3000/login');
            await page.fill('input[name="email"]', 'test@example.com');
            await page.fill('input[name="password"]', 'password123');
            await page.click('button:has-text("Login")');
            await page.waitForURL('**/dashboard');
        }
        """,
        "created_at": datetime.now(timezone.utc),
    }


@pytest.fixture
def sample_embedding_vector():
    """Sample embedding vector (384 dimensions)."""
    return [0.1 + (i * 0.0001) for i in range(384)]


# ========================================================================
# MOCK SERVICE FIXTURES
# ========================================================================

@pytest.fixture
def mock_embeddings_service():
    """Mock embeddings service."""
    with patch("app.services.embeddings.embed_text") as mock_embed:
        mock_embed.return_value = [0.1] * 384
        yield mock_embed


@pytest.fixture
def mock_expansion_service():
    """Mock query expansion service."""
    with patch("app.services.expansion.expand_query", new_callable=AsyncMock) as mock_expand:
        with patch("app.services.expansion.normalize_query", new_callable=AsyncMock) as mock_normalize:
            mock_normalize.return_value = "login test"
            mock_expand.return_value = ["login test", "authentication test", "user login"]
            yield {"expand": mock_expand, "normalize": mock_normalize}


@pytest.fixture
def mock_ranking_service():
    """Mock ranking service."""
    with patch("app.services.ranking.build_candidates") as mock_build:
        with patch("app.services.ranking.select_final_results") as mock_select:
            mock_build.return_value = [
                {"payload": {"_id": "1", "Test Case ID": "TC001"}, "local_score_norm": 0.95},
                {"payload": {"_id": "2", "Test Case ID": "TC002"}, "local_score_norm": 0.87},
            ]
            mock_select.return_value = [
                {"payload": {"_id": "1", "Test Case ID": "TC001"}, "local_score_norm": 0.95},
                {"payload": {"_id": "2", "Test Case ID": "TC002"}, "local_score_norm": 0.87},
            ]
            yield {"build": mock_build, "select": mock_select}


@pytest.fixture
def mock_dedupe_service():
    """Mock deduplication service."""
    with patch("app.services.dedupe_verifier.llm_verify_duplicate", new_callable=AsyncMock) as mock_verify:
        with patch("app.services.dedupe_search_helper.search_similar_testcases", new_callable=AsyncMock) as mock_search:
            with patch("app.services.dedupe_summary.generate_dedupe_summary", new_callable=AsyncMock) as mock_summary:
                mock_verify.return_value = False  # Not a duplicate by default
                mock_search.return_value = []
                mock_summary.return_value = "Test login functionality verification"
                yield {
                    "verify": mock_verify,
                    "search": mock_search,
                    "summary": mock_summary,
                }


@pytest.fixture
def mock_enrichment_service():
    """Mock enrichment service."""
    with patch("app.services.enrichment.get_gemini_enrichment", new_callable=AsyncMock) as mock_enrich:
        mock_enrich.return_value = {
            "summary": "Verify user can login with valid credentials",
            "keywords": ["login", "authentication", "valid", "credentials"],
        }
        yield mock_enrich


@pytest.fixture
def mock_cache_service():
    """Mock cache service."""
    with patch("app.core.cache.get_search_cache", new_callable=AsyncMock) as mock_get:
        with patch("app.core.cache.set_search_cache", new_callable=AsyncMock) as mock_set:
            mock_get.return_value = None
            mock_set.return_value = None
            yield {"get": mock_get, "set": mock_set}


@pytest.fixture
def mock_final_ranking_service():
    """Mock final LLM ranking service."""
    with patch("app.services.finalRanking.final_llm_rerank") as mock_rerank:
        async def rerank_impl(query, results, top_k=None):
            return results
        mock_rerank.side_effect = rerank_impl
        yield mock_rerank


# ========================================================================
# COMPOSITE FIXTURES (common test setups)
# ========================================================================

@pytest.fixture
async def mock_db_with_testcases(
    mock_testcase_collection,
    mock_scripts_collection,
    sample_testcase_doc,
    sample_script_doc,
):
    """Mock database with pre-populated test cases."""
    collection_testcases = mock_testcase_collection
    collection_scripts = mock_scripts_collection
    
    # Setup find_one to return sample data
    collection_testcases.find_one.return_value = sample_testcase_doc
    collection_scripts.find_one.return_value = sample_script_doc
    
    # Setup insert_many to return inserted IDs
    collection_testcases.insert_many.return_value = MagicMock(
        inserted_ids=[sample_testcase_doc["_id"]]
    )
    collection_scripts.insert_many.return_value = MagicMock(
        inserted_ids=[sample_script_doc["_id"]]
    )
    
    return {
        "testcases": collection_testcases,
        "scripts": collection_scripts,
    }


@pytest.fixture
def mock_search_pipeline_mocks(
    mock_embeddings_service,
    mock_expansion_service,
    mock_ranking_service,
    mock_cache_service,
    mock_final_ranking_service,
):
    """All mocks needed for search pipeline."""
    return {
        "embeddings": mock_embeddings_service,
        "expansion": mock_expansion_service,
        "ranking": mock_ranking_service,
        "cache": mock_cache_service,
        "final_ranking": mock_final_ranking_service,
    }


@pytest.fixture
def mock_upload_pipeline_mocks(
    mock_embeddings_service,
    mock_enrichment_service,
    mock_dedupe_service,
):
    """All mocks needed for upload pipeline."""
    return {
        "embeddings": mock_embeddings_service,
        "enrichment": mock_enrichment_service,
        "dedupe": mock_dedupe_service,
    }


# ========================================================================
# HELPER FIXTURES
# ========================================================================

@pytest.fixture
def create_csv_file(tmp_path):
    """Helper to create CSV test files."""
    def _create_csv(filename: str, rows: list[dict]) -> str:
        import csv
        file_path = tmp_path / filename
        
        if not rows:
            file_path.write_text("")
            return str(file_path)
        
        with open(file_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=rows[0].keys())
            writer.writeheader()
            writer.writerows(rows)
        
        return str(file_path)
    
    return _create_csv


@pytest.fixture
def create_excel_file(tmp_path):
    """Helper to create Excel test files."""
    def _create_excel(filename: str, rows: list[dict]) -> str:
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.Workbook()
        ws = wb.active
        
        if not rows:
            wb.save(tmp_path / filename)
            return str(tmp_path / filename)
        
        # Write headers
        for col_idx, header in enumerate(rows[0].keys(), 1):
            ws.cell(row=1, column=col_idx, value=header)
        
        # Write data
        for row_idx, row in enumerate(rows, 2):
            for col_idx, value in enumerate(row.values(), 1):
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        file_path = tmp_path / filename
        wb.save(file_path)
        return str(file_path)
    
    return _create_excel


# ========================================================================
# CLEANUP & MARKERS
# ========================================================================

def pytest_configure(config):
    """Configure pytest with custom markers."""
    config.addinivalue_line("markers", "asyncio: mark test as async")
    config.addinivalue_line("markers", "unit: mark test as unit test")
    config.addinivalue_line("markers", "integration: mark test as integration test")
    config.addinivalue_line("markers", "slow: mark test as slow")
